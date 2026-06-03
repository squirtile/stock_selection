"""Backtest evaluation and Excel report generation."""

from __future__ import annotations

import os
import sys
import time

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(CURRENT_DIR)
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from ml_engine.pattern_extract import extract_indicator_matrix, ML_INDICATOR_COLUMNS, DEFAULT_LOOKBACK
from ml_engine.ml_classifier import MLPatternModel
from ml_engine.similarity import _compute_cosine_similarity_batched



# Excel 报告说明区：写在每个 sheet 最上方，方便打开表格后一眼看懂结果。
def _write_df_with_top_notes(
    writer,
    df: pd.DataFrame,
    sheet_name: str,
    notes: list[str] | None = None,
    index: bool = False,
):
    """Write a DataFrame to Excel with explanatory notes above the table."""
    notes = notes or []
    startrow = len(notes) + 2 if notes else 0
    df.to_excel(writer, sheet_name=sheet_name, index=index, startrow=startrow)

    ws = writer.book[sheet_name]
    max_col = max(1, df.shape[1])

    # 顶部说明区
    if notes:
        for i, text in enumerate(notes, start=1):
            ws.cell(row=i, column=1, value=text)
            ws.cell(row=i, column=1).font = Font(bold=True if i == 1 else False, color="1F2937")
            ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            if max_col > 1:
                ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=max_col)
        blank_row = len(notes) + 1
        ws.cell(row=blank_row, column=1, value="")

    # 表头样式
    header_row = startrow + 1
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="111827")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # 内容样式与列宽
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(header_row, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        # 说明区合并单元格会产生 MergedCell，不能用 ws.columns 的首格取 column_letter。
        # 这里只统计表头和表格内容，避免顶部长说明把列宽撑爆。
        width = min(max(max_len + 2, 10), 28)
        if col_letter == "A":
            width = min(max(width, 12), 20)
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    ws.auto_filter.ref = ws.dimensions


def _similarity_rank_notes() -> list[str]:
    return [
        "【相似度排名说明】本表用于查看候选股票当前最近形态，与模板股票指定强势区间或自动模板之间的相似程度。",
        "平均相似度%：只统计超过阈值的有效匹配结果，再取平均值。它表示这只股票整体有效匹配质量。",
        "最大相似度%：候选股票窗口与所有模板窗口比较时，最高的一次相似度。它表示最像模板的那一次。",
        "匹配次数：相似度达到阈值的次数。次数越多，说明不是偶然只像某一个模板窗口，而是和多个模板阶段都接近。",
        "看表建议：优先关注“最大相似度高 + 匹配次数多”的股票；平均相似度很高但匹配次数只有 1 的股票，需要人工看图确认。",
    ]


def _similarity_detail_notes() -> list[str]:
    return [
        "【明细匹配说明】本表展示每一次候选窗口与模板窗口的具体匹配记录。",
        "候选窗口通常是候选股票当前最近 20 日走势；模板窗口来自手动指定日期区间，或自动识别的启动前/启动期窗口。",
        "可通过模板编号、模板开始/结束日期、候选开始/结束日期，判断股票到底像模板的哪一段。",
    ]


def _template_notes() -> list[str]:
    return [
        "【模板摘要说明】本表记录本次用于匹配的模板窗口来源。",
        "manual/date_range 表示手动指定日期区间；auto_prelaunch 表示自动识别启动点前窗口；recent/recent_fast 表示最近窗口。",
    ]


def _backtest_summary_notes() -> list[str]:
    return [
        "【回测汇总说明】本表统计相似度信号出现后，按不同持有天数计算的整体表现。",
        "回测逻辑：信号日 T 出现相似形态，T+1 开盘买入，持有 N 天后收盘卖出。",
        "重点看：信号次数、胜率%、平均收益率%、最大单笔亏损%、盈亏比。该结果只用于验证形态有效性，不代表实盘收益。",
    ]


def _backtest_detail_notes() -> list[str]:
    return [
        "【回测明细说明】本表列出每一笔历史相似度信号对应的买入、卖出和收益情况。",
        "可用于检查某个股票、某个日期的信号是否真实有效，也可以人工复盘失败案例。",
    ]


def _model_stats_notes() -> list[str]:
    return [
        "【模型统计说明】本表记录机器学习模型训练或验证过程中产生的统计指标。",
        "如果本次只是做相似度匹配，没有训练模型，则本表可能为空或不会生成。",
    ]


def _can_buy_next_day(df: pd.DataFrame, buy_idx: int, skip_limit_up: bool = True) -> bool:
    if not skip_limit_up or buy_idx <= 0 or buy_idx >= len(df):
        return True
    pct = pd.to_numeric(df.iloc[buy_idx].get("涨跌幅", np.nan), errors="coerce")
    # 主板大致涨停约 10%；这里用 9.85 粗略过滤一字/大幅高开不可买情形。
    return not (pd.notna(pct) and pct >= 9.85)


def _calc_return(buy_price: float, sell_price: float, fee_bps: float = 0.0, slippage_bps: float = 0.0) -> float:
    buy = buy_price * (1 + slippage_bps / 10000)
    sell = sell_price * (1 - slippage_bps / 10000)
    ret = (sell / buy - 1) * 100
    ret -= fee_bps / 100  # round-trip fee in bps converted to percent
    return ret


def compute_ml_backtest(
    model: MLPatternModel,
    codes: list[str],
    hold_days_list: list[int] | None = None,
    threshold: float = 0.65,
    min_stock_rows: int = 80,
    fee_bps: float = 0.0,
    slippage_bps: float = 0.0,
    skip_limit_up_next_open: bool = True,
    show_progress: bool = True,
    progress_every: int = 5,
) -> pd.DataFrame:
    if hold_days_list is None:
        hold_days_list = [1, 3, 5, 10]

    trades = []
    total = len(codes)
    start_ts = time.time()
    valid_stocks = 0
    signal_stocks = 0

    for i, code in enumerate(codes, start=1):
        df = extract_indicator_matrix(code, min_rows=min_stock_rows)

        if df.empty:
            if show_progress and (i == 1 or i % progress_every == 0 or i == total):
                elapsed = max(time.time() - start_ts, 0.001)
                speed = i / elapsed
                remain = (total - i) / speed if speed > 0 else 0
                print(
                    f"\r  ML回测进度: {i}/{total} | 当前: {code} | 有效: {valid_stocks} | "
                    f"有信号: {signal_stocks} | 交易: {len(trades)} | 预计剩余: {remain/60:.1f} 分钟",
                    end="",
                    flush=True,
                )
            continue

        indicator_arr = df[model.feature_cols].values
        closes = df["收盘"].values
        opens = df["开盘"].values
        dates = df["日期"].values
        vol_avg_20 = df.get("过去20日平均成交量", pd.Series(np.nan, index=df.index)).values
        volumes = df["成交量"].values
        amounts = df["成交额"].values
        pct_changes = df["涨跌幅"].values
        n = len(df)

        if n < model.lookback + max(hold_days_list):
            continue

        valid_stocks += 1

        # 关键优化：先收集这只股票所有有效窗口，然后一次性批量预测
        window_records = []
        window_vectors = []

        for t in range(model.lookback - 1, n - 1):
            window = indicator_arr[t - model.lookback + 1:t + 1, :]
            if not np.all(np.isfinite(window)):
                continue
            window_records.append(t)
            window_vectors.append(window.flatten().astype(np.float64))

        if not window_vectors:
            continue

        try:
            X = np.array(window_vectors, dtype=np.float64)
            scores = model.predict_proba(X)
        except Exception:
            continue

        stock_trade_start = len(trades)

        for idx, t in enumerate(window_records):
            score = float(scores[idx])
            if score < threshold:
                continue

            for hold_days in hold_days_list:
                buy_idx = t + 1
                sell_idx = t + hold_days

                if buy_idx >= n or sell_idx >= n:
                    continue

                if not _can_buy_next_day(df, buy_idx, skip_limit_up_next_open):
                    continue

                buy_price, sell_price = opens[buy_idx], closes[sell_idx]

                if pd.isna(buy_price) or pd.isna(sell_price) or buy_price <= 0 or sell_price <= 0:
                    continue

                return_pct = _calc_return(float(buy_price), float(sell_price), fee_bps, slippage_bps)
                vol_ratio = volumes[t] / vol_avg_20[t] if pd.notna(vol_avg_20[t]) and vol_avg_20[t] > 0 else None

                trades.append({
                    "代码": code,
                    "信号日期": str(dates[t])[:10],
                    "买入日期": str(dates[buy_idx])[:10],
                    "卖出日期": str(dates[sell_idx])[:10],
                    "买入价": round(float(buy_price), 4),
                    "卖出价": round(float(sell_price), 4),
                    "持有天数": hold_days,
                    "收益率%": round(float(return_pct), 2),
                    "是否盈利": return_pct > 0,
                    "ML分数": round(score, 4),
                    "信号日收盘价": round(float(closes[t]), 4),
                    "信号日涨跌幅": round(float(pct_changes[t]), 2) if pd.notna(pct_changes[t]) else None,
                    "信号日量比": round(float(vol_ratio), 2) if vol_ratio is not None else None,
                    "信号日成交额": float(amounts[t]) if pd.notna(amounts[t]) else None,
                })

        if len(trades) > stock_trade_start:
            signal_stocks += 1

        if show_progress and (i == 1 or i % progress_every == 0 or i == total):
            elapsed = max(time.time() - start_ts, 0.001)
            speed = i / elapsed
            remain = (total - i) / speed if speed > 0 else 0
            print(
                f"\r  ML回测进度: {i}/{total} | 当前: {code} | 有效: {valid_stocks} | "
                f"有信号: {signal_stocks} | 交易: {len(trades)} | 预计剩余: {remain/60:.1f} 分钟",
                end="",
                flush=True,
            )

    if show_progress:
        print()

    return pd.DataFrame(trades)

def _prepare_similarity_scaler(template_vectors: list[np.ndarray], stock_vectors: list[np.ndarray]):
    from sklearn.preprocessing import StandardScaler
    s = StandardScaler()
    s.fit(np.vstack([np.array(template_vectors, dtype=np.float64), np.array(stock_vectors, dtype=np.float64)]))
    return s


def compute_similarity_backtest(
    template_vectors: list[np.ndarray],
    codes: list[str],
    hold_days_list: list[int] | None = None,
    similarity_threshold: float = 0.60,
    scaler=None,
    min_stock_rows: int = 80,
    lookback: int = DEFAULT_LOOKBACK,
    fee_bps: float = 0.0,
    slippage_bps: float = 0.0,
    skip_limit_up_next_open: bool = True,
    show_progress: bool = True,
    progress_every: int = 10,
) -> pd.DataFrame:
    """
    Similarity backtest. Scans every historical window of every code.
    A signal is generated when max similarity >= threshold.
    """
    if hold_days_list is None:
        hold_days_list = [1, 3, 5, 10]
    tpl_matrix_raw = np.array([np.asarray(v, dtype=np.float64) for v in template_vectors], dtype=np.float64)
    trades = []
    total = len(codes)
    start_ts = time.time()
    valid_stocks = 0
    signal_stocks = 0
    for i, code in enumerate(codes, start=1):
        df = extract_indicator_matrix(code, min_rows=min_stock_rows)
        if df.empty or len(df) < lookback + max(hold_days_list):
            if show_progress and (i == 1 or i % progress_every == 0 or i == total):
                elapsed = max(time.time() - start_ts, 0.001)
                speed = i / elapsed
                remain = (total - i) / speed if speed > 0 else 0
                print(f"\r  回测进度: {i}/{total} | 当前: {code} | 有效: {valid_stocks} | 有信号: {signal_stocks} | 交易: {len(trades)} | 预计剩余: {remain/60:.1f} 分钟", end="", flush=True)
            continue
        valid_stocks += 1
        indicator_arr = df[ML_INDICATOR_COLUMNS].values
        closes = df["收盘"].values
        opens = df["开盘"].values
        dates = df["日期"].values
        vol_avg_20 = df.get("过去20日平均成交量", pd.Series(np.nan, index=df.index)).values
        volumes = df["成交量"].values
        amounts = df["成交额"].values
        pct_changes = df["涨跌幅"].values
        n = len(df)

        stock_records = []
        stock_vectors = []
        for t in range(lookback - 1, n - 1):
            window = indicator_arr[t - lookback + 1:t + 1, :]
            if np.all(np.isfinite(window)):
                stock_records.append(t)
                stock_vectors.append(window.flatten().astype(np.float64))
        if not stock_vectors:
            continue

        s = scaler or _prepare_similarity_scaler(template_vectors, stock_vectors)
        tpl_scaled = s.transform(tpl_matrix_raw)
        stock_scaled = s.transform(np.array(stock_vectors, dtype=np.float64))

        # For each stock vector compute max similarity to any template.
        max_sims = np.zeros(len(stock_scaled), dtype=np.float64)
        for tpl in tpl_scaled:
            sims = _compute_cosine_similarity_batched(tpl, stock_scaled)
            max_sims = np.maximum(max_sims, sims)

        stock_trade_start = len(trades)
        for idx, t in enumerate(stock_records):
            max_sim = float(max_sims[idx])
            if max_sim < similarity_threshold:
                continue
            for hold_days in hold_days_list:
                buy_idx = t + 1
                sell_idx = t + hold_days
                if buy_idx >= n or sell_idx >= n:
                    continue
                if not _can_buy_next_day(df, buy_idx, skip_limit_up_next_open):
                    continue
                buy_price, sell_price = opens[buy_idx], closes[sell_idx]
                if pd.isna(buy_price) or pd.isna(sell_price) or buy_price <= 0 or sell_price <= 0:
                    continue
                return_pct = _calc_return(float(buy_price), float(sell_price), fee_bps, slippage_bps)
                vol_ratio = volumes[t] / vol_avg_20[t] if pd.notna(vol_avg_20[t]) and vol_avg_20[t] > 0 else None
                trades.append({
                    "代码": code,
                    "信号日期": str(dates[t])[:10],
                    "买入日期": str(dates[buy_idx])[:10],
                    "卖出日期": str(dates[sell_idx])[:10],
                    "买入价": round(float(buy_price), 4),
                    "卖出价": round(float(sell_price), 4),
                    "持有天数": hold_days,
                    "收益率%": round(float(return_pct), 2),
                    "是否盈利": return_pct > 0,
                    "相似度": round(max_sim, 4),
                    "相似度%": round(max_sim * 100, 1),
                    "信号日收盘价": round(float(closes[t]), 4),
                    "信号日涨跌幅": round(float(pct_changes[t]), 2) if pd.notna(pct_changes[t]) else None,
                    "信号日量比": round(float(vol_ratio), 2) if vol_ratio is not None else None,
                    "信号日成交额": float(amounts[t]) if pd.notna(amounts[t]) else None,
                })
        if len(trades) > stock_trade_start:
            signal_stocks += 1
        if show_progress and (i == 1 or i % progress_every == 0 or i == total):
            elapsed = max(time.time() - start_ts, 0.001)
            speed = i / elapsed
            remain = (total - i) / speed if speed > 0 else 0
            print(f"\r  回测进度: {i}/{total} | 当前: {code} | 有效: {valid_stocks} | 有信号: {signal_stocks} | 交易: {len(trades)} | 预计剩余: {remain/60:.1f} 分钟", end="", flush=True)
    if show_progress:
        print()
    return pd.DataFrame(trades)


def summarize_ml_backtest(results_df: pd.DataFrame, hold_days: int | None = None) -> pd.DataFrame:
    if results_df.empty:
        return pd.DataFrame()
    df = results_df.copy()
    if hold_days is not None:
        df = df[df["持有天数"] == hold_days]
    if df.empty:
        return pd.DataFrame()
    wins = df[df["是否盈利"] == True]
    losses = df[df["是否盈利"] == False]
    avg_loss = abs(losses["收益率%"].mean()) if len(losses) else 0.0
    avg_win = wins["收益率%"].mean() if len(wins) else 0.0
    pl_ratio = avg_win / avg_loss if avg_loss > 0 else None
    return pd.DataFrame([{
        "持有天数": hold_days if hold_days is not None else "all",
        "信号次数": len(df),
        "盈利次数": len(wins),
        "亏损次数": len(losses),
        "胜率%": round(len(wins) / len(df) * 100, 2),
        "平均收益率%": round(df["收益率%"].mean(), 2),
        "中位数收益率%": round(df["收益率%"].median(), 2),
        "最大单笔收益%": round(df["收益率%"].max(), 2),
        "最大单笔亏损%": round(df["收益率%"].min(), 2),
        "平均盈利%": round(avg_win, 2),
        "平均亏损%": round(avg_loss, 2),
        "盈亏比": round(pl_ratio, 2) if pl_ratio else None,
    }])


def summarize_ml_by_hold_days(results_df: pd.DataFrame) -> pd.DataFrame:
    if results_df.empty:
        return pd.DataFrame()
    rows = []
    for hd in sorted(results_df["持有天数"].unique()):
        s = summarize_ml_backtest(results_df, int(hd))
        if not s.empty:
            rows.append(s)
    return pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()


def generate_similarity_report(
    similarity_detail: pd.DataFrame,
    similarity_stock: pd.DataFrame,
    model_stats: dict | None = None,
    template_info: list[dict] | None = None,
    backtest_trades: pd.DataFrame | None = None,
    backtest_summary: pd.DataFrame | None = None,
    output_file: str = "output/ml_similarity/report.xlsx",
):
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        if similarity_stock is not None and not similarity_stock.empty:
            _write_df_with_top_notes(
                writer,
                similarity_stock,
                sheet_name="相似度排名",
                notes=_similarity_rank_notes(),
                index=False,
            )
        if similarity_detail is not None and not similarity_detail.empty:
            _write_df_with_top_notes(
                writer,
                similarity_detail,
                sheet_name="明细匹配",
                notes=_similarity_detail_notes(),
                index=False,
            )
        if template_info:
            _write_df_with_top_notes(
                writer,
                pd.DataFrame(template_info),
                sheet_name="模板摘要",
                notes=_template_notes(),
                index=False,
            )
        if backtest_summary is not None and not backtest_summary.empty:
            _write_df_with_top_notes(
                writer,
                backtest_summary,
                sheet_name="回测汇总",
                notes=_backtest_summary_notes(),
                index=False,
            )
        if backtest_trades is not None and not backtest_trades.empty:
            _write_df_with_top_notes(
                writer,
                backtest_trades,
                sheet_name="回测明细",
                notes=_backtest_detail_notes(),
                index=False,
            )
        if model_stats:
            rows = []
            for k, v in model_stats.items():
                if isinstance(v, dict):
                    for kk, vv in v.items():
                        rows.append({"指标": f"{k}/{kk}", "值": vv})
                else:
                    rows.append({"指标": k, "值": v})
            _write_df_with_top_notes(
                writer,
                pd.DataFrame(rows),
                sheet_name="模型统计",
                notes=_model_stats_notes(),
                index=False,
            )
    return output_file
